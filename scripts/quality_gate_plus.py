#!/usr/bin/env python3
"""Enhanced quality gate for mathematical modeling projects.

Unlike the lightweight 02_代码/06_quality_gate.py, this script checks whether
key artifacts are not only present but also minimally substantive:
- result tables are readable and non-empty;
- baseline/core/sensitivity evidence exists;
- report text exists and appears to answer problem sections;
- figure/table references are present when corresponding assets exist;
- state machine and project_meta are consistent;
- final package and pipeline summary are visible when expected.

Creates:
  06_过程记录/质量门禁/quality_gate_plus.json
  06_过程记录/质量门禁/quality_gate_plus.md

This is a workflow integrity gate, not a proof of mathematical correctness.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

REPORT_EXTS = {".md", ".tex", ".docx", ".pdf"}
TABLE_EXTS = {".csv", ".json", ".xlsx", ".xls"}
FIGURE_EXTS = {".png", ".jpg", ".jpeg", ".svg", ".pdf"}


@dataclass
class Finding:
    item: str
    status: str  # pass/warn/fail
    detail: str
    evidence: str = ""


def find_files(project: Path, rel: str, exts: set[str] | None = None) -> list[Path]:
    root = project / rel
    if not root.exists():
        return []
    files = [p for p in root.rglob("*") if p.is_file() and not p.name.startswith("~$")]
    if exts is not None:
        files = [p for p in files if p.suffix.lower() in exts]
    return sorted(files)


def rel(project: Path, paths: list[Path], limit: int = 20) -> str:
    return "; ".join(str(p.relative_to(project)) for p in paths[:limit])


def read_text_safe(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return ""


def docx_text(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as zf:
            xml = zf.read("word/document.xml")
        root = ET.fromstring(xml)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        return "\n".join(node.text or "" for node in root.findall(".//w:t", ns))
    except Exception:
        return ""


def report_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".md", ".tex"}:
        return read_text_safe(path)
    if suffix == ".docx":
        return docx_text(path)
    return ""


def table_profile(path: Path) -> dict:
    suffix = path.suffix.lower()
    profile = {"path": str(path), "rows": 0, "cols": 0, "readable": False, "error": ""}
    try:
        if suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as f:
                rows = list(csv.reader(f))
            profile["rows"] = max(0, len(rows) - 1) if rows else 0
            profile["cols"] = max((len(r) for r in rows), default=0)
            profile["readable"] = True
        elif suffix == ".json":
            obj = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(obj, list):
                profile["rows"] = len(obj)
                profile["cols"] = len(obj[0]) if obj and isinstance(obj[0], dict) else 1
            elif isinstance(obj, dict):
                profile["rows"] = len(obj)
                profile["cols"] = 2
            else:
                profile["rows"] = 1
                profile["cols"] = 1
            profile["readable"] = True
        elif suffix in {".xlsx", ".xls"}:
            try:
                import openpyxl  # type: ignore
            except Exception as e:
                profile["error"] = f"openpyxl unavailable: {e!r}"
                return profile
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            rows = 0
            cols = 0
            for ws in wb.worksheets:
                rows += max(0, (ws.max_row or 0) - 1)
                cols = max(cols, ws.max_column or 0)
            wb.close()
            profile.update({"rows": rows, "cols": cols, "readable": True})
    except Exception as e:
        profile["error"] = repr(e)
    return profile


def has_named_result(files: list[Path], patterns: list[str]) -> list[Path]:
    out = []
    for p in files:
        name = p.name.lower()
        if any(re.search(pat, name, flags=re.I) for pat in patterns):
            out.append(p)
    return sorted(out)


def substantive_problem_analysis(project: Path) -> bool:
    path = project / "06_过程记录" / "problem_analysis.md"
    text = read_text_safe(path)
    stripped = re.sub(r"[#\s`|:\-/\[\]（）()]+", "", text)
    placeholders = ["待补充", "todo", "TODO", "待完成"]
    if any(p in text for p in placeholders) and len(stripped) < 120:
        return False
    headings = ["小问", "数据", "输出", "题型", "目标"]
    return len(stripped) >= 60 and sum(1 for h in headings if h in text) >= 2


def parse_state_markdown(project: Path) -> str:
    path = project / "06_过程记录" / "状态机" / "PROJECT_STATE.md"
    text = read_text_safe(path)
    m = re.search(r"当前连续完成状态：`?(S-?\d+)`?", text)
    return m.group(1) if m else ""


def audit(project: Path, expect_final: bool = False) -> dict:
    findings: list[Finding] = []

    standard_dirs = [
        "00_题目与资料", "01_原始数据", "02_代码", "03_结果表格", "04_图表",
        "05_报告定稿", "06_过程记录", "07_提交包",
    ]
    missing_dirs = [d for d in standard_dirs if not (project / d).exists()]
    findings.append(Finding("standard_directories", "fail" if missing_dirs else "pass", f"missing={len(missing_dirs)}", "; ".join(missing_dirs)))

    materials = find_files(project, "00_题目与资料", None) + find_files(project, "01_原始数据", None)
    findings.append(Finding("materials_or_raw_data", "pass" if materials else "warn", f"found={len(materials)}", rel(project, materials)))

    findings.append(Finding("problem_analysis_substantive", "pass" if substantive_problem_analysis(project) else "warn", "problem_analysis.md 是否有实质题解信息", "06_过程记录/problem_analysis.md"))

    tables = find_files(project, "03_结果表格", TABLE_EXTS)
    table_profiles = [table_profile(p) for p in tables]
    unreadable = [p for p, prof in zip(tables, table_profiles) if not prof["readable"]]
    empty = [p for p, prof in zip(tables, table_profiles) if prof["readable"] and (prof["rows"] <= 0 or prof["cols"] <= 0)]
    findings.append(Finding("result_tables_readable", "fail" if unreadable else ("pass" if tables else "warn"), f"tables={len(tables)}, unreadable={len(unreadable)}", rel(project, unreadable)))
    findings.append(Finding("result_tables_nonempty", "fail" if empty else ("pass" if tables else "warn"), f"empty={len(empty)}", rel(project, empty)))

    baseline = has_named_result(tables, [r"baseline", r"基线"])
    core = has_named_result(tables, [r"model_results", r"main_model", r"core_model", r"主模型", r"核心模型"])
    sensitivity = has_named_result(tables, [r"sensitivity", r"robust", r"敏感", r"鲁棒"])
    findings.append(Finding("baseline_result_exists", "pass" if baseline else "fail", f"found={len(baseline)}", rel(project, baseline)))
    findings.append(Finding("core_model_result_exists", "pass" if core else "fail", f"found={len(core)}", rel(project, core)))
    findings.append(Finding("sensitivity_result_exists", "pass" if sensitivity else "warn", f"found={len(sensitivity)}", rel(project, sensitivity)))

    figures = find_files(project, "04_图表", FIGURE_EXTS)
    tiny_figs = [p for p in figures if p.stat().st_size < 16]
    findings.append(Finding("figures_exist", "pass" if figures else "warn", f"figures={len(figures)}", rel(project, figures)))
    findings.append(Finding("figures_not_tiny", "warn" if tiny_figs else "pass", f"tiny={len(tiny_figs)}", rel(project, tiny_figs)))

    reports = find_files(project, "05_报告定稿", REPORT_EXTS)
    texts = [report_text(p) for p in reports]
    combined = "\n".join(t for t in texts if t)
    findings.append(Finding("report_exists", "pass" if reports else "fail", f"reports={len(reports)}", rel(project, reports)))
    findings.append(Finding("report_text_substantive", "pass" if len(re.sub(r"\s+", "", combined)) >= 120 else "warn", f"text_chars={len(combined)}", "PDF 不做文本深度解析" if reports and not combined else ""))
    question_hits = len(re.findall(r"(?:问题|小问|第[一二三四五六七八九十1-9]\s*[问题]?|Question\s*\d+)", combined, flags=re.I))
    findings.append(Finding("report_answers_problem_sections", "pass" if question_hits >= 1 else "warn", f"question_section_hits={question_hits}", "报告应逐问给出结论"))
    fig_ref_hits = len(re.findall(r"(?:图\s*\d+|Figure\s*\d+|\.(?:png|jpg|jpeg|svg))", combined, flags=re.I))
    table_ref_hits = len(re.findall(r"(?:表\s*\d+|Table\s*\d+|\.(?:csv|xlsx|xls|json))", combined, flags=re.I))
    findings.append(Finding("report_references_figures", "pass" if (not figures or fig_ref_hits > 0) else "warn", f"figure_refs={fig_ref_hits}, figures={len(figures)}"))
    findings.append(Finding("report_references_tables", "pass" if (not tables or table_ref_hits > 0) else "warn", f"table_refs={table_ref_hits}, tables={len(tables)}"))

    audit_json = project / "06_过程记录" / "一致性检查" / "auto_report_audit.json"
    if audit_json.exists():
        audit_data = json.loads(audit_json.read_text(encoding="utf-8"))
        counts = audit_data.get("counts", {})
        status = "fail" if counts.get("fail", 0) else ("warn" if counts.get("warn", 0) else "pass")
        findings.append(Finding("auto_report_audit_clean", status, f"pass={counts.get('pass', 'NA')} warn={counts.get('warn', 'NA')} fail={counts.get('fail', 'NA')}", str(audit_json.relative_to(project))))
    else:
        findings.append(Finding("auto_report_audit_clean", "warn", "auto_report_audit.json 不存在", str(audit_json.relative_to(project))))

    meta_path = project / "project_meta.json"
    meta = json.loads(meta_path.read_text(encoding="utf-8")) if meta_path.exists() else {}
    meta_state = meta.get("highest_contiguous_state", "")
    md_state = parse_state_markdown(project)
    if meta_state and md_state:
        status = "pass" if meta_state == md_state else "fail"
        detail = f"meta={meta_state}, markdown={md_state}"
    else:
        status = "warn"
        detail = f"meta={meta_state or 'missing'}, markdown={md_state or 'missing'}"
    findings.append(Finding("state_meta_consistent", status, detail, "project_meta.json; 06_过程记录/状态机/PROJECT_STATE.md"))

    final_dir = project / "07_提交包"
    final_required = [final_dir / "README_submit.md", final_dir / "SHA256SUMS.txt", final_dir / "submission_manifest.json"]
    missing_final = [p for p in final_required if not p.exists()]
    final_status = "fail" if expect_final and missing_final else ("pass" if not missing_final else "warn")
    findings.append(Finding("final_package_complete", final_status, f"missing={len(missing_final)}", "; ".join(str(p.relative_to(project)) for p in missing_final)))

    pipeline_summary = project / "06_过程记录" / "pipeline" / "pipeline_run_summary.json"
    findings.append(Finding("pipeline_summary_exists", "pass" if pipeline_summary.exists() else ("warn" if expect_final else "pass"), str(pipeline_summary.relative_to(project))))

    counts = {
        "findings": len(findings),
        "pass": sum(1 for f in findings if f.status == "pass"),
        "warn": sum(1 for f in findings if f.status == "warn"),
        "fail": sum(1 for f in findings if f.status == "fail"),
    }
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "project": str(project),
        "expect_final": expect_final,
        "counts": counts,
        "table_profiles": table_profiles,
        "findings": [asdict(f) for f in findings],
    }


def write_outputs(project: Path, summary: dict) -> tuple[Path, Path]:
    out_dir = project / "06_过程记录" / "质量门禁"
    out_dir.mkdir(parents=True, exist_ok=True)
    json_path = out_dir / "quality_gate_plus.json"
    md_path = out_dir / "quality_gate_plus.md"
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    lines: list[str] = []
    lines.append("# 增强质量门禁报告\n\n")
    lines.append(f"生成时间：{summary['generated_at']}\n\n")
    lines.append(f"项目：`{summary['project']}`\n\n")
    c = summary["counts"]
    lines.append(f"## 汇总\n\n- pass: {c['pass']}\n- warn: {c['warn']}\n- fail: {c['fail']}\n\n")
    lines.append("## 检查项\n\n| 项目 | 状态 | 说明 | 证据 |\n|---|---|---|---|\n")
    for f in summary["findings"]:
        mark = {"pass": "✅", "warn": "⚠️", "fail": "❌"}.get(f["status"], f["status"])
        detail = str(f["detail"]).replace("|", "\\|")
        evidence = str(f.get("evidence", "")).replace("|", "\\|")
        lines.append(f"| {f['item']} | {mark} {f['status']} | {detail} | {evidence} |\n")
    lines.append("\n## 结果表读取概况\n\n")
    lines.append("| 文件 | readable | rows | cols | error |\n|---|---:|---:|---:|---|\n")
    for p in summary.get("table_profiles", []):
        err = str(p.get("error", "")).replace("|", "\\|")
        lines.append(f"| `{Path(p['path']).name}` | {p['readable']} | {p['rows']} | {p['cols']} | {err} |\n")
    md_path.write_text("".join(lines), encoding="utf-8")
    return json_path, md_path


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("project")
    parser.add_argument("--strict", action="store_true", help="Return non-zero on warnings as well as failures")
    parser.add_argument("--expect-final", action="store_true", help="Require final package and pipeline summary evidence")
    args = parser.parse_args()

    project = Path(args.project).expanduser().resolve()
    if not project.exists():
        print(f"Project not found: {project}", file=sys.stderr)
        return 2
    summary = audit(project, expect_final=args.expect_final)
    json_path, md_path = write_outputs(project, summary)
    c = summary["counts"]
    print(f"Quality gate plus: {md_path}")
    print(f"Quality gate plus json: {json_path}")
    print(f"Findings: {c['findings']} total, {c['fail']} fail, {c['warn']} warn")
    if c["fail"]:
        return 1
    if args.strict and c["warn"]:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
